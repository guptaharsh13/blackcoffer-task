import csv
from openpyxl import load_workbook
# from scraper_selenium import findContent
# bs4 would be much faster in the current situation (tested) - compares by a time difference of 16 minutes
from scraper_bs4 import findContent
import os
from pathlib import Path
from data_analysis import analyseFiles
import concurrent.futures
from logger_config import configLogger

main_logger = configLogger(__name__)


def readCSV(filename):
    with open(file=filename, newline="") as csvfile:
        reader = csv.DictReader(csvfile)
        urls = {}
        for row in reader:
            url_id = row.get("URL_ID")
            url = row.get("URL")
            if url_id and url:
                urls[url_id] = url
    return urls


def readXLSX(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.active

    urls = {}

    for row in sheet.iter_rows(min_row=2, max_row=171, min_col=1, max_col=2, values_only=True):
        url_id = row[0]
        url = row[1]
        urls[url_id] = url
    return urls


def saveContent(url_id, content, dir_name):
    dir_path = os.path.join(Path(__file__).resolve(
    ).parent.parent, dir_name)

    if not os.path.exists(dir_path):
        os.makedirs(dir_path)

    with open(file=f"{dir_path}/{url_id}.txt", mode="w") as file:
        file.write(content)


def combine(url_id, url, dir_name):
    if os.path.exists(os.path.join(Path(__file__).resolve().parent.parent, f"{dir_name}/{url_id}.txt")):
        return
    title, content = findContent(url=url)
    saveContent(
        url_id=url_id, content=f"{title.strip()}\n\n{content.strip()}", dir_name=dir_name)


def main():

    input_file_path = os.path.join(Path(__file__).resolve(
    ).parent.parent, "Input.csv")

    # read from xlsx is slower as compared to csv, Thus, I have made use of a csv file over an xlsx file to read input data (url_ids and urls)
    # urls = readXLSX(input_file_path)
    urls = readCSV(input_file_path)

    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = [executor.submit(combine, url_id, url, "articles")
                   for url_id, url in urls.items()]

        done, not_done = concurrent.futures.wait(
            fs=futures, return_when=concurrent.futures.ALL_COMPLETED)
        if not_done:
            main_logger.critical("Unable to scrape data from URLs")
            quit({"error": "Unable to scrape data from URLs"})
        if done:
            main_logger.info("completed URLs extraction")

    analyseFiles(dir_name="articles", urls=urls)


if __name__ == "__main__":
    main()
